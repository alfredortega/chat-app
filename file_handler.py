"""
file_handler.py — upload storage, text extraction, and size validation.
"""

import csv
import os
import time

# ── Configuration ──────────────────────────────────────────────────────────────

UPLOAD_ROOT = os.path.join(os.path.dirname(__file__), "uploads")

# Warn the user if total extracted text exceeds this many characters (~40k tokens)
WARN_THRESHOLD = 150_000

# Hard-truncate injected content at this limit to avoid oversized API requests
HARD_LIMIT = 500_000

# Cache TTL for linked‑folder scans (seconds)
CACHE_TTL = 60

# Simple in‑process cache for linked‑folder scans: {folder_path: (file_entries, timestamp)}
_linked_folder_cache: dict = {}

ALLOWED_EXTENSIONS = {
    ".txt", ".md", ".markdown",
    ".py", ".js", ".ts", ".jsx", ".tsx",
    ".html", ".htm", ".css", ".scss",
    ".json", ".yaml", ".yml", ".toml", ".ini", ".env",
    ".sh", ".bat", ".ps1",
    ".sql",
    ".csv",
    ".pdf",
    ".docx",
    ".xlsx", ".xls",
    ".xml",
    ".log",
    ".rs", ".go", ".java", ".c", ".cpp", ".h", ".cs", ".rb", ".php",
    # Image formats
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp",
}


def ensure_upload_dir(conv_id: int) -> str:
    """Return (and create if needed) the upload directory for a conversation."""
    path = os.path.join(UPLOAD_ROOT, str(conv_id))
    os.makedirs(path, exist_ok=True)
    return path


def allowed_extension(filename: str) -> bool:
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def extract_text(filepath: str, original_filename: str) -> tuple[str, bool]:
    """
    Extract plain text from a file.
    Returns (text, was_truncated).
    """
    ext = os.path.splitext(original_filename)[1].lower()

    try:
        if ext == ".pdf":
            text = _extract_pdf(filepath)
        elif ext == ".docx":
            text = _extract_docx(filepath)
        elif ext in (".xlsx", ".xls"):
            text = _extract_xlsx(filepath)
        elif ext == ".csv":
            text = _extract_csv(filepath)
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp"):
            # Image files – no textual content extracted.
            text = ""
        else:
            # Plain text / code / markup
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
    except Exception as exc:
        return f"[Could not extract content: {exc}]", False

    truncated = False
    if len(text) > HARD_LIMIT:
        text = text[:HARD_LIMIT]
        truncated = True

    return text, truncated


# Fix #7: removed unused char_count() — callers use len(extract_text(...)[0]) directly.


# ── Format helpers ─────────────────────────────────────────────────────────────

def _extract_pdf(filepath: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(filepath)
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n\n".join(pages)


def _extract_docx(filepath: str) -> str:
    from docx import Document
    doc = Document(filepath)
    return "\n".join(para.text for para in doc.paragraphs)


def _extract_xlsx(filepath: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            rows.append(" | ".join("" if cell is None else str(cell) for cell in row))
        if rows:
            sheets.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets)


def _extract_csv(filepath: str) -> str:
    rows = []
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(", ".join(row))
    return "\n".join(rows)


# ── Context injection ──────────────────────────────────────────────────────────

def build_file_context(files: list[dict]) -> str:
    """
    Given a list of uploaded file dicts (from DB), return a formatted string to
    inject into the system prompt so the model can reference the files.
    """
    if not files:
        return ""

    parts = ["The following files have been uploaded for this conversation and are available for reference:\n"]
    for f in files:
        filepath = f["disk_path"]
        filename = f["original_name"]
        text, truncated = extract_text(filepath, filename)
        header = f"--- FILE: {filename} ---"
        footer = "--- END OF FILE ---"
        if truncated:
            footer = f"--- END OF FILE (truncated at {HARD_LIMIT:,} characters) ---"
        parts.append(f"{header}\n{text}\n{footer}")

    return "\n\n".join(parts)


# ── Linked folder scanning ─────────────────────────────────────────────────────

def scan_linked_folder(folder_path: str) -> list[dict]:
    """
    Return a list of dicts for every linkable file at ``folder_path``.

    ``folder_path`` may point at either a directory or a single file:
      • Directory — walked recursively; every file with an allowed
        extension is returned.
      • File — returned as a single entry if its extension is allowed.

    Each dict has the shape:
        { "abs_path": str, "rel_path": str, "filename": str }
    Sorted by relative path for deterministic ordering.
    """
    results = []
    # Check cache
    if folder_path in _linked_folder_cache:
        cached_entries, cached_ts = _linked_folder_cache[folder_path]
        if time.time() - cached_ts < CACHE_TTL:
            return cached_entries

    folder_path = os.path.normpath(folder_path)

# Single-file link
    if os.path.isfile(folder_path):
        fname = os.path.basename(folder_path)
        ext = os.path.splitext(fname)[1].lower()
        if ext in ALLOWED_EXTENSIONS:
            results.append({
                "abs_path": folder_path,
                "rel_path": fname,
                "filename": fname,
            })
        # Store in cache before returning
        _linked_folder_cache[folder_path] = (results, time.time())
        return results

    if not os.path.isdir(folder_path):
        return results

    for root, dirs, files in os.walk(folder_path):
        # Skip hidden directories (e.g. .git, .venv)
        dirs[:] = [d for d in sorted(dirs) if not d.startswith(".")]
        for fname in sorted(files):
            if fname.startswith("."):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ALLOWED_EXTENSIONS:
                continue
            abs_path = os.path.join(root, fname)
            rel_path = os.path.relpath(abs_path, folder_path).replace("\\", "/")
            results.append({
                "abs_path": abs_path,
                "rel_path": rel_path,
                "filename": fname,
            })
    # Store in cache before returning
    _linked_folder_cache[folder_path] = (results, time.time())
    return results


def build_linked_folder_context(
    linked_folders: list[dict],
    uploaded_files: list[dict],
) -> tuple[str, int]:
    """
    Read all files from linked folders/files live from disk.
    Each linked entry may point to a directory or an individual file.
    Linked files take priority over uploaded files with the same basename.
    Returns (context_str, total_char_count).
    """
    if not linked_folders:
        return "", 0

    # Build a set of basenames already covered by linked folders (for dedup)
    linked_basenames: set[str] = set()
    folder_sections: list[str] = []
    total_chars = 0

    for lf in linked_folders:
        folder_path = lf["folder_path"]
        is_file      = os.path.isfile(folder_path)
        label        = "LINKED FILE" if is_file else "LINKED FOLDER"
        file_entries = scan_linked_folder(folder_path)

        if not file_entries:
            folder_sections.append(
                f"--- {label}: {folder_path} (no supported files found) ---"
            )
            continue

        if is_file:
            # A single linked file — emit it directly without a folder wrapper.
            entry = file_entries[0]
            linked_basenames.add(entry["filename"].lower())
            text, truncated = extract_text(entry["abs_path"], entry["filename"])
            total_chars += len(text)
            footer = "--- END OF FILE ---"
            if truncated:
                footer = f"--- END OF FILE (truncated at {HARD_LIMIT:,} chars) ---"
            folder_sections.append(
                f"--- LINKED FILE: {folder_path} ---\n{text}\n{footer}"
            )
            continue

        parts = [f"--- LINKED FOLDER: {folder_path} ({len(file_entries)} files) ---\n"]
        for entry in file_entries:
            linked_basenames.add(entry["filename"].lower())
            text, truncated = extract_text(entry["abs_path"], entry["filename"])
            total_chars += len(text)
            header = f"  -- FILE: {entry['rel_path']} --"
            footer = "  -- END OF FILE --"
            if truncated:
                footer = f"  -- END OF FILE (truncated at {HARD_LIMIT:,} chars) --"
            parts.append(f"{header}\n{text}\n{footer}")

        folder_sections.append("\n\n".join(parts))

    context = (
        "The following linked folders and files are available for reference "
        "(content is read live from disk):\n\n"
        + "\n\n".join(folder_sections)
    )

    # Append uploaded files that are NOT superseded by a linked-folder file
    remaining_uploads = [
        f for f in uploaded_files
        if os.path.basename(f["original_name"]).lower() not in linked_basenames
    ]
    if remaining_uploads:
        upload_ctx = build_file_context(remaining_uploads)
        if upload_ctx:
            context += "\n\n" + upload_ctx

    return context, total_chars
